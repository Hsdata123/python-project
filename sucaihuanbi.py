import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from data_functions import process_columns
from sucai import live_room
# 读取Excel文件
file_path = f'合并后_消耗前20素材_{live_room}.xlsx'
df = pd.read_excel(file_path)

# 确保日期列是日期类型
df['投放日期'] = pd.to_datetime(df['投放日期'])
df = process_columns(df)
df.to_excel("111.xlsx",index=False)
# 选择需要计算环比的指标列
metrics_columns = [
    '整体展现次数', '整体点击次数', '整体点击率', '整体转化率', '整体消耗', 
    '基础消耗', '整体支付ROI', '整体成交金额', '整体成交订单数', '整体成交订单成本',
    '用户实际支付金额', '整体千次展现费用', '视频点赞数', '新增粉丝数', '平均观看时长',
    '视频播放数', '视频完播率', '视频评论数', '2秒播放率', '3秒播放率', '5秒播放率', '10秒播放率'
]

# 按日期汇总数据
daily_summary = df.groupby('投放日期')[metrics_columns].sum().reset_index()

# 计算数值型指标的环比（百分比变化）
numeric_columns = [col for col in metrics_columns if col not in ['整体点击率', '整体转化率', '视频完播率', '2秒播放率', '3秒播放率', '5秒播放率', '10秒播放率']]
rate_columns = ['整体点击率', '整体转化率', '视频完播率', '2秒播放率', '3秒播放率', '5秒播放率', '10秒播放率']

# 按日期排序
daily_summary = daily_summary.sort_values('投放日期')

# 计算环比
daily_summary_环比 = daily_summary.copy()

for col in numeric_columns:
    daily_summary_环比[f'{col}_环比'] = daily_summary[col].pct_change() * 100  # 转换为百分比

for col in rate_columns:
    # 对于比率指标，我们需要先计算总体的比率，而不是比率的和
    # 这里我们暂时用平均值的变化来近似表示环比
    daily_summary_环比[f'{col}_环比'] = daily_summary[col].pct_change() * 100

# 添加日期说明列
daily_summary_环比['日期'] = daily_summary_环比['投放日期'].dt.strftime('%Y-%m-%d')
daily_summary_环比['环比说明'] = daily_summary_环比['投放日期'].dt.strftime('%Y-%m-%d') + ' vs ' + daily_summary_环比['投放日期'].shift(1).dt.strftime('%Y-%m-%d')

# 重新排列列的顺序，将日期相关列放在前面
cols = ['投放日期', '日期', '环比说明'] + [col for col in daily_summary_环比.columns if col not in ['投放日期', '日期', '环比说明']]
daily_summary_环比 = daily_summary_环比[cols]

# 创建一个更简洁的环比报告，只保留环比列
环比报告列 = ['投放日期', '日期', '环比说明']
for col in metrics_columns:
    环比报告列.append(f'{col}_环比')

环比报告 = daily_summary_环比[环比报告列].copy()

# 格式化环比值为百分比字符串
for col in 环比报告.columns:
    if '_环比' in col:
        环比报告[col] = 环比报告[col].apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "N/A")

# 使用openpyxl将结果写入原Excel文件的新工作表中
try:
    # 尝试加载现有工作簿
    book = load_workbook(file_path)
    
    # 删除已存在的环比工作表（如果存在）
    if '日期汇总环比' in book.sheetnames:
        del book['日期汇总环比']
    
    # 创建新工作表
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    writer.book = book
    
    # 将环比报告写入新工作表
    环比报告.to_excel(writer, sheet_name='日期汇总环比', index=False)
    
    # 调整列宽
    worksheet = writer.sheets['日期汇总环比']
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    writer.save()
    print("环比报告已成功写入原Excel文件的新工作表中")
    
except Exception as e:
    print(f"写入文件时出错: {e}")
    # 如果无法修改原文件，则创建新文件
    new_file_path = f'合并后_消耗前20素材_{live_room}_含环比.xlsx'
    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='原始数据', index=False)
        环比报告.to_excel(writer, sheet_name='日期汇总环比', index=False)
    print(f"已创建新文件: {new_file_path}")

# 显示环比报告的前几行
print("\n日期汇总环比报告预览:")
print(环比报告.head())