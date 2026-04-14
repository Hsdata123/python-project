from datetime import datetime
import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime,timedelta
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def column_index_from_string(column_letter):
    """将列字母转换为数字索引 (A=1, B=2, ..., Z=26, AA=27, etc.)"""
    column_letter = column_letter.upper()
    index = 0
    for char in column_letter:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def write_dataframe_to_excel_template(template_path, output_path, df_dict, 
                                     start_cell_dict=None, 
                                     write_headers_dict=None):
    """
    将多个DataFrame写入Excel模板的指定位置
    
    参数:
    - template_path: 模板文件路径
    - output_path: 输出文件路径
    - df_dict: 字典，格式为 {sheet_name: dataframe} 或 {sheet_name: (dataframe, start_cell, write_headers)}
    - start_cell_dict: 可选，字典，格式为 {sheet_name: start_cell}
    - write_headers_dict: 可选，字典，格式为 {sheet_name: write_headers}
    """
    try:
        # 检查模板文件是否存在
        if not os.path.exists(template_path):
            print(f"错误: 模板文件不存在: {template_path}")
            return False
        
        # 加载模板工作簿
        workbook = load_workbook(template_path)
        
        # 处理每个sheet的写入
        success_count = 0
        total_sheets = len(df_dict)
        
        for sheet_name, df_info in df_dict.items():
            try:
                # 解析参数
                if isinstance(df_info, tuple) and len(df_info) >= 1:
                    df = df_info[0]
                    start_cell = df_info[1] if len(df_info) > 1 else (
                        start_cell_dict.get(sheet_name, 'A1') if start_cell_dict else 'A1'
                    )
                    write_headers = df_info[2] if len(df_info) > 2 else (
                        write_headers_dict.get(sheet_name, True) if write_headers_dict else True
                    )
                else:
                    df = df_info
                    start_cell = (
                        start_cell_dict.get(sheet_name, 'A1') if start_cell_dict else 'A1'
                    )
                    write_headers = (
                        write_headers_dict.get(sheet_name, True) if write_headers_dict else True
                    )
                
                # 检查工作表是否存在
                if sheet_name not in workbook.sheetnames:
                    print(f"警告: 工作表 '{sheet_name}' 不存在于模板中，将创建新工作表")
                    worksheet = workbook.create_sheet(sheet_name)
                else:
                    worksheet = workbook[sheet_name]
                
                # 解析起始单元格
                start_col_letter = ''.join(filter(str.isalpha, start_cell))
                start_row = int(''.join(filter(str.isdigit, start_cell)))
                
                # 将列字母转换为数字索引
                start_col_idx = column_index_from_string(start_col_letter)
                
                # 将DataFrame转换为行数据
                rows = list(dataframe_to_rows(df, index=False, header=write_headers))
                
                # 遍历行和单元格写入数据
                for r_idx, row in enumerate(rows):
                    for c_idx, value in enumerate(row):
                        # 计算实际单元格位置
                        cell_row = start_row + r_idx
                        cell_col = start_col_idx + c_idx
                        
                        # 写入值
                        worksheet.cell(row=cell_row, column=cell_col, value=value)
                
                print(f"✓ 工作表 '{sheet_name}': 成功写入 {len(df)} 行数据 (起始位置: {start_cell})")
                success_count += 1
                
            except Exception as e:
                print(f"✗ 工作表 '{sheet_name}' 写入失败: {str(e)}")
                continue
        
        # 保存工作簿
        workbook.save(output_path)
        print(f"\n写入完成: {success_count}/{total_sheets} 个工作表写入成功")
        print(f"输出文件: {output_path}")
        
        return success_count > 0
        
    except Exception as e:
        print(f"写入模板时发生错误: {str(e)}")
        return False

def process_columns(df):
    df_processed = df.copy()
    
    for col in df_processed.columns:
        if '率' in col:
            # 处理百分比列
            df_processed.loc[:, col] = (
                df_processed[col].astype(str)
                .apply(lambda x: str(float(x.replace('%', '')) / 100) if '%' in x else x)
                .replace(['nan', ''], '0')
                .astype(float)
            )
        else :
            try:
            # 处理金额列
                df_processed.loc[:, col] = (
                    df_processed[col].astype(str)
                    .str.replace(r'[¥,]', '', regex=True)
                    .replace(['nan', ''], '0')
                    .astype(float)
                )
            except:
                pass
    return df_processed

def merge_excel_files(folder_path, file_name_pattern):
    """
    合并多个相同字段的Excel和CSV文件到一个DataFrame
    
    参数:
    folder_path (str): 文件夹路径
    file_name_pattern (str): 文件名模式，支持通配符
    
    返回:
    pd.DataFrame: 合并后的DataFrame
    """
    # 构建Excel和CSV文件的路径模式
    excel_pattern = os.path.join(folder_path, f"{file_name_pattern}.xlsx")
    csv_pattern = os.path.join(folder_path, f"{file_name_pattern}.csv")
    
    # 获取匹配的文件列表
    excel_files = glob.glob(excel_pattern)
    csv_files = glob.glob(csv_pattern)
    file_list = excel_files + csv_files
    
    # 如果没有找到文件，返回空的DataFrame
    if not file_list:
        print(f"警告: 在路径 {folder_path} 下没有找到匹配 {file_name_pattern} 的Excel或CSV文件")
        return pd.DataFrame()
    
    # 存储所有DataFrame的列表
    dfs = []
    
    # 遍历所有文件
    for file_path in file_list:
        try:
            # 根据文件扩展名选择读取方式
            file_extension = os.path.splitext(file_path)[1].lower()
            
            if file_extension == '.xlsx':
                df = pd.read_excel(file_path)
            elif file_extension == '.csv':
                # 自动检测编码格式
                encodings = ['utf-8', 'gbk', 'gb2312', 'latin1']
                df = None
                for encoding in encodings:
                    try:
                        df = pd.read_csv(file_path, encoding=encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                if df is None:
                    raise Exception("无法确定文件编码")
            else:
                print(f"跳过不支持的文件格式: {file_path}")
                continue
            
            # 添加文件名作为一列，便于追踪数据来源
            df['source_file'] = os.path.basename(file_path)
            
            # 添加到列表
            dfs.append(df)
            
            print(f"成功读取: {os.path.basename(file_path)}，包含 {len(df)} 行数据")
            
        except Exception as e:
            print(f"读取文件 {file_path} 时出错: {str(e)}")
            continue
    
    # 合并所有DataFrame
    if dfs:
        merged_df = pd.concat(dfs, ignore_index=True)
        print(f"合并完成！总共合并了 {len(file_list)} 个文件，总计 {len(merged_df)} 行数据")
        return merged_df
    else:
        print("没有成功读取任何文件")
        return pd.DataFrame()
    
def get_product_name(name_series):
    # 将系列转换为字符串类型
    name_str = name_series.astype(str)
    
    # 统计包含"椰子"的数量
    coconut_count = name_str.str.contains("椰子").sum()
    
    # 统计包含"鱼子酱"的数量
    caviar_count = name_str.str.contains("鱼子酱").sum()
    
    # 比较数量并返回结果
    if coconut_count > caviar_count:
        return "椰子"
    elif caviar_count > coconut_count:
        return "鱼子酱"
    else:
        # 如果数量相等或都为0，返回"其他"
        return "其他"
    

def get_year_month_week_v2(date_str):
    """
    输入日期字符串（如"2024-05-15"），返回"年-月-当月第几周"格式
    规则：从1号开始算，到下一个周日为第1周结束，后续每周为周一至周日
    """
    # 转换输入日期为datetime对象
    date = datetime.strptime(date_str, "%Y-%m-%d")
    year = date.year
    month = date.month
    day = date.day

    # 1. 找到当月1号
    first_day = datetime(year, month, 1)
    # 2. 找到当月第一个周日（isoweekday()：1=周一，7=周日）
    first_day_weekday = first_day.isoweekday()
    # 计算1号到首个周日的天数：若1号是周日（7），则0天；否则7 - 1号的周几
    days_to_first_sunday = 7 - first_day_weekday if first_day_weekday != 7 else 0
    first_sunday = first_day + timedelta(days=days_to_first_sunday)

    # 3. 判断目标日期所在周数
    if date <= first_sunday:
        # 日期在1号至首个周日之间 → 第1周
        week_number = 1
    else:
        # 日期在首个周日之后：计算首个周日到目标日期的间隔，再按7天/周划分
        days_after_first_sunday = (date - first_sunday).days
        # 间隔天数 // 7 + 2（因首个周日之后的周一为第2周开始）
        week_number = (days_after_first_sunday // 7) + 2

    return f"{year}-{month:02d}-第{week_number}周"


def get_previous_sunday(date=None):
    """
    计算给定日期的上一个周日
    
    Args:
        date: 日期，可以是datetime对象或字符串（YYYY-MM-DD），默认为今天
    
    Returns:
        datetime: 上一个周日的日期
    """
    if date is None:
        date = datetime.now()
    elif isinstance(date, str):
        date = datetime.strptime(date, "%Y-%m-%d")
    
    # 获取当前日期的星期几（周一=0, 周日=6）
    current_weekday = date.weekday()
    
    # 计算到上一个周日的差值
    # 如果今天是周日，则上一个周日是7天前
    # 如果今天是其他日子，计算 days_back = (current_weekday + 1) % 7 + 7
    days_back = (current_weekday + 1) % 7 + 7
    
    previous_sunday = date - timedelta(days=days_back)
    return previous_sunday.strftime('%Y-%m-%d')

def get_previous_previous_monday(date=None):
    """
    计算给定日期的上上个周一
    
    Args:
        date: 日期，可以是datetime对象或字符串（YYYY-MM-DD），默认为今天
    
    Returns:
        datetime: 上上个周一的日期
    """
    if date is None:
        date = datetime.now()
    elif isinstance(date, str):
        date = datetime.strptime(date, "%Y-%m-%d")
    
    # 先计算上一个周一，再减去7天得到上上个周一
    current_weekday = date.weekday()
    days_to_previous_monday = current_weekday + 7
    previous_previous_monday = date - timedelta(days=days_to_previous_monday + 7)
    
    return previous_previous_monday.strftime('%Y-%m-%d')

def get_last_two_complete_weeks(date=None):
    """
    获取最近两个完整周的日期范围
    
    Returns:
        dict: 包含两个完整周信息的字典
    """
    if date is None:
        date = datetime.now()
    elif isinstance(date, str):
        date = datetime.strptime(date, "%Y-%m-%d")
    
    # 计算上一个周日（最近完整周的结束）
    prev_sunday = get_previous_sunday(date)
    
    # 计算上一个周一（最近完整周的开始）
    prev_monday = prev_sunday - timedelta(days=6)
    
    # 计算上上个周一（上上个完整周的开始）
    prev_prev_monday = get_previous_previous_monday(date)
    
    # 计算上上个周日（上上个完整周的结束）
    prev_prev_sunday = prev_prev_monday + timedelta(days=6)
    
    return {
        '最近完整周': {
            '周一': prev_monday.strftime('%Y-%m-%d'),
            '周日': prev_sunday.strftime('%Y-%m-%d')
        },
        '上上个完整周': {
            '周一': prev_prev_monday.strftime('%Y-%m-%d'),
            '周日': prev_prev_sunday.strftime('%Y-%m-%d')
        }
    }